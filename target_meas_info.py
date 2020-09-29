import os
import glob
import openpyxl
import xmltodict
from config import *
from meas_param_scrape import *

#col_string_names , col_key_list

#The path to the goods you want to scrape
file_path = "C://Python3//Scripts//CD_SEM_XML//device_cd_recipes//4IU8000AC//"

wb_save_file_path ="Target_param.xlsx"

scraped_meas_params_dict = scrape_meas_params(file_path + "MeasParams//")




target_xml_file_paths = glob.glob(file_path+"Target//" + "*CCD*.xml")


#Create XL workbook and name the first sheet.
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Targets"

#Create the Column names
col = 1
for tic in col_string_names:
    sheet.cell(row=1,column = col).value = tic
    col += 1
sheet.cell(row=1,column = col).value = "Meas_param_Process_directory"
col += 1
sheet.cell(row=1,column = col).value = "Meas_param_Product_directory"
col += 1
sheet.cell(row=1,column = col).value = "Meas_param_Layer_directory"
col += 1
sheet.cell(row=1,column = col).value = "Meas_param_Name"
col += 1
for meas_param_key_list in meas_params:
    for meas_param_keys in meas_param_key_list:
        sheet.cell(row=1,column = col).value = meas_param_keys
        col += 1
sheet.cell(row=1,column = col).value = "Meas_param_id_Verify"
row = 2

for xml_file in target_xml_file_paths:
    col = 1
    doc = get_xml_dict(xml_file)
    for column_key in col_key_list:
        val = get_xml_val(doc, column_key)
        try:
            val =round(float(val),4)
            sheet.cell(row=row,column = col).value = val
        except ValueError:
            sheet.cell(row=row,column = col).value = val
        col += 1
    current_meas_param_list = scraped_meas_params_dict[get_xml_val(doc,['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '2', 'Value', 'References', 'Reference', 'Value', 'DirectMapping', 'Field', '0', 'Value'])]
    sheet.cell(row=row,column = col).value = current_meas_param_list[0]
    col += 1
    sheet.cell(row=row,column = col).value = current_meas_param_list[1]
    col += 1
    sheet.cell(row=row,column = col).value = current_meas_param_list[2]
    col += 1
    sheet.cell(row=row,column = col).value = current_meas_param_list[3]
    col += 1
    for meas_param_key_list in meas_params:
        for meas_param_keys in meas_param_key_list:
            sheet.cell(row=row,column = col).value = round(float(current_meas_param_list[4][meas_param_keys]), 3 )
            col += 1
    sheet.cell(row=row,column = col).value = get_xml_val(doc,['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '2', 'Value', 'References', 'Reference', 'Value', 'DirectMapping', 'Field', '0', 'Value'])
    row += 1

#Name of the file
wb_save_file_path ="Target_param.xlsx"
wb.save(wb_save_file_path)
print('done bitch')
