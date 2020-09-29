import os
import glob
import openpyxl
import xmltodict
from config import *
from meas_param_scrape import *
from target_param_scrape import *

######==== EDIT THESE VALUES =========

#name of xl doc
wb_save_file_path ="9007_CDrecipe.xlsx"

#file path to the goods. 
file_path = "C://Python3//Scripts//CD_SEM_XML//FNDRY_TO_PROD//5CC900XAC//5CC9007AC//"

#====================================


#key= meas param id, output: [ppl[0], ppl[1], ppl[2], ppl[3], meas_dict]
scraped_meas_params_dict = scrape_meas_params(file_path + "MeasParams//")

#key=target id, output: [ppl[0], ppl[1], ppl[2], ppl[3], target_dict]
scraped_target_params_dict = scrape_target_params(file_path + "Target//")

def recipe_target_info(xml_file):
    target_list = []

    with open(xml_file) as fd:
        doc = xmltodict.parse(fd.read())

    target_dict_key = ['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '1', 'Value', 'One2ManyMapping', 'O2M', '1', 'O2MElement', 'Value', 'References', 'Reference', 'Value', 'DirectMapping', 'Field', '1', 'Value']
    target_id_key =   ['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '1', 'Value', 'One2ManyMapping', 'O2M', '1', 'O2MElement', 'Value', 'References', 'Reference', 'Value', 'DirectMapping', 'Field', '0', 'Value']
    target_x_loc_key = ['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '1', 'Value', 'One2ManyMapping', 'O2M', '1', 'O2MElement', 'Value', 'DirectMapping', 'Field', '6', 'Value']
    target_y_loc_key = ['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '1', 'Value', 'One2ManyMapping', 'O2M', '1', 'O2MElement', 'Value', 'DirectMapping', 'Field', '7', 'Value']
    try:
        target = get_xml_val(doc, target_dict_key)
        target_id = get_xml_val(doc, target_id_key)
        target_x_loc = get_xml_val(doc, target_x_loc_key)
        target_y_loc = get_xml_val(doc, target_y_loc_key)

        target_list.append([target,target_id,target_x_loc,target_y_loc])
    except TypeError:
        target_dict_key.insert(-8,"0")
        target_id_key.insert(-8,"0")
        target_x_loc_key.insert(-5,"0")
        target_y_loc_key.insert(-5,"0")

        t1 = ['Descriptor', 'Value', 'One2OneMapping', 'O2O', '0', 'Value', 'One2OneMapping', 'O2O', '1', 'Value', 'One2ManyMapping', 'O2M', '1', 'O2MElement']
        target_xml_list = get_xml_val(doc, t1)
        for i in range(len(target_xml_list)):
            target_dict_key[-9] = str(i)
            target_id_key[-9] = str(i)
            target_x_loc_key[-6] = str(i)
            target_y_loc_key[-6] = str(i)
            target = get_xml_val(doc, target_dict_key)
            target_id = get_xml_val(doc, target_id_key)
            target_x_loc = get_xml_val(doc, target_x_loc_key)
            target_y_loc = get_xml_val(doc, target_y_loc_key)
            target_list.append([target,target_id,target_x_loc,target_y_loc])
    ppl = os.path.basename(xml_file)[:-4].split("__")
    return ppl,target_list

#create Xl sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "da_info"
col = 1
sheet.cell(row=1, column = col).value = "Recipe Name"
col += 1
target_name_col_list = ["Target Process","Target Product","Target Layer","Target Name"]
for i in target_name_col_list:
    sheet.cell(row=1, column = col).value = i
    col += 1
sheet.cell(row=1, column = col).value = "Target X Loc"
col += 1
sheet.cell(row=1, column = col).value = "Target Y Loc"
col += 1
for i in col_string_names:
    sheet.cell(row=1, column = col).value = i
    col += 1
for i in meas_param_key_list_simple:
    sheet.cell(row=1, column = col).value = i
    col += 1


recipe_file_paths = glob.glob(file_path +"CDRecipe//" + "*.xml")

recipe_list = []
for recipe_file in recipe_file_paths:
    ppl,target_list = recipe_target_info(recipe_file)
    recipe_list.append([ppl[-1],target_list])

print(recipe_list)
row = 2

for recipe_name, recipe_target_list in recipe_list:
    for target_name, target_id,target_x_loc,target_y_loc in recipe_target_list:
        if not "OBS" in target_name:
            col = 1
            sheet.cell(row=row, column = col).value = recipe_name
            col += 1
            print(target_name, target_id)
            #key=target id, output: [ppl[0], ppl[1], ppl[2], ppl[3], target_dict]
            recipe_target = scraped_target_params_dict[target_id]
            sheet.cell(row=row, column = col).value = recipe_target[0]
            col += 1
            sheet.cell(row=row, column = col).value = recipe_target[1]
            col += 1
            sheet.cell(row=row, column = col).value = recipe_target[2]
            col += 1
            sheet.cell(row=row, column = col).value = recipe_target[3]
            col += 1
            sheet.cell(row=row, column = col).value = round(float(target_x_loc),3)
            col += 1
            sheet.cell(row=row, column = col).value = round(float(target_y_loc),3)
            col += 1

            for target_col_name in col_string_names:
                sheet.cell(row=row, column = col).value = recipe_target[4][target_col_name]
                col += 1
            #"meas_param_id"
            target_meas_params = scraped_meas_params_dict[recipe_target[4]["meas_param_id"]]
            for meas_param_col_name in meas_param_key_list:
                sheet.cell(row=row, column = col).value = round(float(target_meas_params[4][meas_param_col_name]),3)
                col += 1
            row += 1



wb.save(wb_save_file_path)
print('done bitch')













#
